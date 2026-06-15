"""
Fits three model families and writes results to output/OLS_Model_Results.xlsx:

  1. LEVELS    - log(real price) on log(price)_{t-1} plus macro factors.
                 Segmented 4 ways: {Light, Heavy} x {short-war, long-war}.
                 R^2 is inflated because the lag dominates; kept for back-compat.
  2. RETURNS   - dlog(real price) on dlog(macro) plus event dummies.
                 Lower R^2 but unbiased. Same 4-way segmentation.
  3. DIFFERENTIAL - (WCS - WTI) in $/bbl, single model. The Enbridge-relevant
                 spec: heavy/light spread driven by pipeline takeaway and US
                 storage dynamics.

All models use Newey-West HAC standard errors (maxlags=6).
"""
from __future__ import annotations

import numpy as np
import pandas as pd
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.stats.diagnostic import acorr_ljungbox
from statsmodels.tsa.stattools import coint, adfuller
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import config

# (crude_label, B1, dep_levels, lag_levels, dep_returns, lag_returns)
CRUDE_SPECS = [
    ("Light", 1, "log_wti_real", "log_wti_lag1", "dlog_wti_real", "dlog_wti_lag1"),
    ("Heavy", 0, "log_wcs_real", "log_wcs_lag1", "dlog_wcs_real", "dlog_wcs_lag1"),
]
WAR_SPECS = [("Short-term", 0), ("Long-term", 1)]


def _resolve(regressors: list[str], placeholder: str, replacement: str) -> list[str]:
    return [replacement if x == placeholder else x for x in regressors]


def _resolve_levels(lag_var: str) -> list[str]:
    return _resolve(config.REGRESSORS, "log_price_lag1", lag_var)


def _resolve_returns(lag_var: str) -> list[str]:
    return _resolve(config.REGRESSORS_RETURNS, "dlog_price_lag1", lag_var)


def _ensure_lag_returns(df: pd.DataFrame, lag_var: str) -> pd.DataFrame:
    """Compute dlog_*_lag1 on the fly so the returns spec doesn't need
    pre-baked columns."""
    base = lag_var.replace("_lag1", "")
    if lag_var not in df.columns and base in df.columns:
        df = df.copy()
        df[lag_var] = df[base].shift(1)
    return df


def _drop_constant_cols(sub: pd.DataFrame, regressors: list[str]) -> list[str]:
    """Drop regressors that are constant within the fit sub-sample (perfectly
    collinear with the intercept). Common when a regime dummy has 0 (or 1)
    coverage in a particular war-state slice."""
    keep = []
    for r in regressors:
        if sub[r].nunique(dropna=True) > 1:
            keep.append(r)
    return keep


def _fit(df: pd.DataFrame, dep: str, regressors: list[str]):
    cols = [dep] + regressors
    sub = df[cols].dropna()
    regressors = _drop_constant_cols(sub, regressors)
    if len(sub) < len(regressors) + 5:
        raise ValueError(f"Too few observations: {len(sub)} (need {len(regressors)+5}+)")
    X = sm.add_constant(sub[regressors])
    y = sub[dep]
    return sm.OLS(y, X).fit(cov_type="HAC", cov_kwds={"maxlags": 6}), sub


def _coef_table(res) -> pd.DataFrame:
    return pd.DataFrame({
        "coef":      res.params,
        "std_err":   res.bse,
        "t":         res.tvalues,
        "p":         res.pvalues,
        "ci_low":    res.conf_int()[0],
        "ci_high":   res.conf_int()[1],
    }).round(5)


def _diag_table(res, n_obs: int) -> pd.DataFrame:
    """Diagnostics including Ljung-Box test for residual autocorrelation.

    Ljung-Box (lag=12): tests whether the residuals contain ANY autocorrelation
    out to 12 lags. P-value < 0.05 means residuals are NOT white noise --
    the model has missed some dynamics (e.g. a missing AR term, regime, or
    seasonality).  P-value > 0.05 means residuals look like white noise:
    the model has captured what's predictable.

    Durbin-Watson tests only first-order serial correlation; Ljung-Box is the
    more rigorous, multi-lag analogue used in the time-series literature.
    """
    lb = acorr_ljungbox(res.resid, lags=[12], return_df=True)
    lb_stat = float(lb["lb_stat"].iloc[0]); lb_p = float(lb["lb_pvalue"].iloc[0])
    return pd.DataFrame({
        "metric": ["n_obs", "R2", "adj_R2", "F-stat", "F p-value",
                   "AIC", "BIC", "Durbin-Watson",
                   "Ljung-Box(12) stat", "Ljung-Box(12) p-value"],
        "value":  [n_obs, res.rsquared, res.rsquared_adj, res.fvalue, res.f_pvalue,
                   res.aic, res.bic, sm.stats.stattools.durbin_watson(res.resid),
                   lb_stat, lb_p],
    }).round(4)


def _cointegration_table(df: pd.DataFrame) -> pd.DataFrame:
    """Engle-Granger cointegration test on log(WTI_real) and log(WCS_real).

    Two prices that drift together but have a stationary spread are
    *cointegrated*. If so, modelling them as two independent OLS regressions
    is mis-specified: a Vector Error Correction Model (VECM) is the correct
    spec, since it imposes the long-run equilibrium relationship.

    We also run ADF tests on each series and on the spread, plus a stationarity
    check on the WCS-WTI differential (which is what our DIFFERENTIAL model
    relies on -- if the spread is non-stationary, that model is also mis-specified).
    """
    rows = []

    def _adf(name: str, s: pd.Series):
        x = s.dropna()
        if len(x) < 24:
            rows.append({"test": f"ADF({name})", "statistic": float("nan"),
                         "p_value": float("nan"), "n": len(x),
                         "interpretation": "too few obs"})
            return
        stat, p, *_ = adfuller(x, autolag="AIC")
        rows.append({"test": f"ADF({name})", "statistic": round(stat, 3),
                     "p_value": round(p, 4), "n": len(x),
                     "interpretation": "stationary (reject unit root)" if p < 0.05
                     else "non-stationary (has unit root)"})

    # ADFs
    for col, name in [("log_wti_real", "log WTI"),
                      ("log_wcs_real", "log WCS"),
                      ("wcs_wti_diff", "WCS-WTI diff")]:
        if col in df.columns:
            _adf(name, df[col])

    # Engle-Granger: regress y on x, then ADF on residuals.
    # statsmodels.tsa.stattools.coint returns the t-stat and p-value directly.
    s1 = df["log_wti_real"].dropna()
    s2 = df["log_wcs_real"].dropna()
    common = s1.index.intersection(s2.index)
    if len(common) >= 30:
        try:
            stat, p, _crit = coint(s1.loc[common], s2.loc[common], trend="c")
            rows.append({"test": "Engle-Granger coint(log WTI, log WCS)",
                         "statistic": round(float(stat), 3),
                         "p_value": round(float(p), 4),
                         "n": len(common),
                         "interpretation": "cointegrated (reject H0 of no-coint)"
                         if p < 0.05
                         else "NOT cointegrated at 5% (cannot reject H0)"})
        except Exception as e:
            rows.append({"test": "Engle-Granger coint(log WTI, log WCS)",
                         "statistic": float("nan"), "p_value": float("nan"),
                         "n": len(common), "interpretation": f"failed: {e}"})
    return pd.DataFrame(rows)


def _vif_table(df: pd.DataFrame, dep: str, regressors: list[str]) -> pd.DataFrame:
    sub = df[[dep] + regressors].dropna()
    X = sm.add_constant(sub[regressors])
    rows = []
    for i, name in enumerate(X.columns):
        if name == "const":
            continue
        try:
            v = variance_inflation_factor(X.values, i)
        except Exception:
            v = float("nan")
        rows.append({
            "variable": name,
            "VIF": round(v, 3),
            "flag": "HIGH" if v > 10 else ("watch" if v > 5 else ""),
        })
    return pd.DataFrame(rows)


def _write_sheet(wb: Workbook, name: str, header: dict, coef_df: pd.DataFrame,
                 diag_df: pd.DataFrame, vif_df: pd.DataFrame | None,
                 resid_df: pd.DataFrame | None) -> None:
    ws = wb.create_sheet(name[:31])
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DCE6F1")

    ws.append([f"Model: {name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Specification"])
    ws["A3"].font = bold
    for k, v in header.items():
        ws.append([k, v])
    ws.append([])

    ws.append(["Coefficients"])
    ws.cell(row=ws.max_row, column=1).font = bold
    coef_df = coef_df.reset_index().rename(columns={"index": "variable"})
    for r in dataframe_to_rows(coef_df, index=False, header=True):
        ws.append(r)
    for cell in ws[ws.max_row - len(coef_df)]:
        cell.font = bold; cell.fill = hdr_fill
    ws.append([])

    ws.append(["Diagnostics"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for r in dataframe_to_rows(diag_df, index=False, header=True):
        ws.append(r)
    for cell in ws[ws.max_row - len(diag_df)]:
        cell.font = bold; cell.fill = hdr_fill

    if vif_df is not None and len(vif_df):
        ws.append([])
        ws.append(["Multicollinearity (VIF)"])
        ws.cell(row=ws.max_row, column=1).font = bold
        for r in dataframe_to_rows(vif_df, index=False, header=True):
            ws.append(r)
        for cell in ws[ws.max_row - len(vif_df)]:
            cell.font = bold; cell.fill = hdr_fill

    if resid_df is not None and len(resid_df):
        ws.append([])
        ws.append(["Residuals (sample)"])
        ws.cell(row=ws.max_row, column=1).font = bold
        for r in dataframe_to_rows(resid_df.tail(24).reset_index(), index=False, header=True):
            ws.append(r)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 28)


def _fit_levels_family(df: pd.DataFrame, wb: Workbook, summary_rows: list):
    print("\n[ LEVELS family — log(real price) ]")
    for crude_label, b1, dep_lvl, lag_lvl, _, _ in CRUDE_SPECS:
        for war_label, b2 in WAR_SPECS:
            sub = df[df["war_long"] == b2]
            regs = _resolve_levels(lag_lvl)
            # Filter to columns that exist (Canadian production is optional)
            regs = [r for r in regs if r in df.columns]
            try:
                res, fit_df = _fit(sub, dep_lvl, regs)
            except ValueError as e:
                print(f"  [SKIP] {crude_label} | {war_label}: {e}")
                continue
            n_obs = int(res.nobs)
            sheet = f"L_{crude_label[:1]}_{war_label.replace('-', '')[:5]}"  # e.g. L_L_Short
            header = {
                "Family": "LEVELS  (log price ~ macro factors)",
                "Crude (B1)":   f"{crude_label} ({b1})",
                "War (B2)":     f"{war_label} ({b2})",
                "Dependent":    dep_lvl,
                "Sample n":     n_obs,
                "Sample range": f"{fit_df.index.min().date()} to {fit_df.index.max().date()}",
                "SE method":    "Newey-West HAC, maxlags=6",
            }
            coef = _coef_table(res)
            diag = _diag_table(res, n_obs)
            vif  = _vif_table(sub, dep_lvl, regs)
            resid = pd.DataFrame({"actual": res.fittedvalues + res.resid,
                                  "fitted": res.fittedvalues,
                                  "residual": res.resid})
            _write_sheet(wb, sheet, header, coef, diag, vif, resid)
            summary_rows.append({
                "Family": "Levels", "Sheet": sheet,
                "Crude": crude_label, "War": war_label,
                "n": n_obs, "R2": round(res.rsquared, 4),
                "adj_R2": round(res.rsquared_adj, 4),
                "DW": round(sm.stats.stattools.durbin_watson(res.resid), 3),
            })
            print(f"  [OK] {crude_label:5s} | {war_label:10s}  n={n_obs:3d}  R2={res.rsquared:.3f}")


def _fit_returns_family(df: pd.DataFrame, wb: Workbook, summary_rows: list):
    print("\n[ RETURNS family - dlog(real price) ]")
    for crude_label, b1, _, _, dep_ret, lag_ret in CRUDE_SPECS:
        df_ext = _ensure_lag_returns(df, lag_ret)
        for war_label, b2 in WAR_SPECS:
            sub = df_ext[df_ext["war_long"] == b2]
            regs = _resolve_returns(lag_ret)
            # Filter to columns that exist (some optional like SPR may be missing)
            regs = [r for r in regs if r in df_ext.columns]
            try:
                res, fit_df = _fit(sub, dep_ret, regs)
            except ValueError as e:
                print(f"  [SKIP] {crude_label} | {war_label}: {e}")
                continue
            n_obs = int(res.nobs)
            sheet = f"R_{crude_label[:1]}_{war_label.replace('-', '')[:5]}"
            header = {
                "Family": "RETURNS  (dlog price ~ dfactors + events)",
                "Crude (B1)":   f"{crude_label} ({b1})",
                "War (B2)":     f"{war_label} ({b2})",
                "Dependent":    dep_ret,
                "Sample n":     n_obs,
                "Sample range": f"{fit_df.index.min().date()} to {fit_df.index.max().date()}",
                "SE method":    "Newey-West HAC, maxlags=6",
                "Note": "R^2 will be much lower than the levels model — this is correct. Returns are mostly noise.",
            }
            coef = _coef_table(res)
            diag = _diag_table(res, n_obs)
            vif  = _vif_table(sub, dep_ret, regs)
            resid = pd.DataFrame({"actual": res.fittedvalues + res.resid,
                                  "fitted": res.fittedvalues,
                                  "residual": res.resid})
            _write_sheet(wb, sheet, header, coef, diag, vif, resid)
            summary_rows.append({
                "Family": "Returns", "Sheet": sheet,
                "Crude": crude_label, "War": war_label,
                "n": n_obs, "R2": round(res.rsquared, 4),
                "adj_R2": round(res.rsquared_adj, 4),
                "DW": round(sm.stats.stattools.durbin_watson(res.resid), 3),
            })
            print(f"  [OK] {crude_label:5s} | {war_label:10s}  n={n_obs:3d}  R2={res.rsquared:.3f}  (lower is honest)")


def _fit_differential(df: pd.DataFrame, wb: Workbook, summary_rows: list):
    print("\n[ DIFFERENTIAL — WCS minus WTI ($/bbl), single model ]")
    regs = list(config.REGRESSORS_DIFFERENTIAL)
    if "apportionment_lag1" in df.columns and df["apportionment_lag1"].notna().any():
        regs.append("apportionment_lag1")
        print(f"  apportionment_lag1 included ({df['apportionment_lag1'].notna().sum()} obs)")
    regs = [r for r in regs if r in df.columns]
    try:
        res, fit_df = _fit(df, "wcs_wti_diff", regs)
    except ValueError as e:
        print(f"  [SKIP] {e}")
        return
    n_obs = int(res.nobs)
    sheet = "D_WCS_WTI_diff"
    header = {
        "Family": "DIFFERENTIAL  (heavy-light spread)",
        "Dependent":    "wcs_wti_diff (WCS - WTI, $/bbl, real)",
        "Sample n":     n_obs,
        "Sample range": f"{fit_df.index.min().date()} to {fit_df.index.max().date()}",
        "SE method":    "Newey-West HAC, maxlags=6",
        "Use":          "Models the WCS-WTI spread directly. Most relevant for Enbridge mainline economics.",
    }
    coef = _coef_table(res)
    diag = _diag_table(res, n_obs)
    vif  = _vif_table(df, "wcs_wti_diff", regs)
    resid = pd.DataFrame({"actual": res.fittedvalues + res.resid,
                          "fitted": res.fittedvalues,
                          "residual": res.resid})
    _write_sheet(wb, sheet, header, coef, diag, vif, resid)
    summary_rows.append({
        "Family": "Differential", "Sheet": sheet,
        "Crude": "WCS-WTI", "War": "(all)",
        "n": n_obs, "R2": round(res.rsquared, 4),
        "adj_R2": round(res.rsquared_adj, 4),
        "DW": round(sm.stats.stattools.durbin_watson(res.resid), 3),
    })
    print(f"  [OK] WCS-WTI diff  n={n_obs:3d}  R2={res.rsquared:.3f}")


def _write_cointegration_sheet(df: pd.DataFrame, wb: Workbook) -> None:
    """Add a Cointegration sheet to the workbook -- ADF and Engle-Granger tests."""
    print("\n[ COINTEGRATION DIAGNOSTICS ]")
    tab = _cointegration_table(df)
    if tab.empty:
        print("  [SKIP] insufficient data"); return
    ws = wb.create_sheet("Cointegration")
    ws.append(["Stationarity and Cointegration tests"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    notes = [
        "ADF tests each log-price series and the WCS-WTI spread for stationarity.",
        "  p-value < 0.05  => stationary (no unit root).",
        "  p-value > 0.05  => non-stationary (random-walk-like).",
        "Engle-Granger tests whether log WTI and log WCS share a stationary",
        "long-run relationship (cointegration).",
        "  p < 0.05  => cointegrated => a VECM is theoretically the correct spec",
        "              (separate OLS regressions on the two prices is mis-specified).",
        "  p > 0.05  => not cointegrated at 5%; OLS spec is acceptable.",
        "The DIFFERENTIAL model REQUIRES the spread to be stationary -- a",
        "non-stationary spread would mean the differential model is unreliable.",
    ]
    for n in notes:
        ws.append([n])
    ws.append([])
    ws.append(["Test results"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for r in dataframe_to_rows(tab, index=False, header=True):
        ws.append(r)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col), default=10
        ) + 2
    for _, row in tab.iterrows():
        print(f"  {row['test']:50s}  stat={row['statistic']}  p={row['p_value']}  -> {row['interpretation']}")


def _write_lasso_sheet(df: pd.DataFrame, wb: Workbook) -> None:
    """Lasso feature pruning on the RETURNS spec for each crude.

    Uses LassoCV (cross-validated alpha) to identify which regressors actually
    add signal. Variables with zero coefficient after Lasso are candidates to
    prune. Standardised features so coefficients are scale-comparable.

    This sheet is *advisory* -- the production model continues to use the full
    spec for backward compatibility, but the sheet shows which variables Lasso
    would keep and which it would drop.
    """
    try:
        from sklearn.linear_model import LassoCV
        from sklearn.preprocessing import StandardScaler
    except ImportError:
        print("\n[ LASSO ] scikit-learn not installed; skipping feature-pruning sheet")
        return

    print("\n[ LASSO feature pruning (advisory) ]")
    rows = []
    for crude_label, _b1, _dep_lvl, _lag_lvl, dep_ret, lag_ret in CRUDE_SPECS:
        df_ext = _ensure_lag_returns(df, lag_ret)
        regs = _resolve_returns(lag_ret)
        regs = [r for r in regs if r in df_ext.columns]
        sub = df_ext[[dep_ret] + regs].dropna()
        if len(sub) < 40:
            print(f"  [SKIP] {crude_label}: too few obs ({len(sub)})")
            continue
        X = sub[regs].values
        y = sub[dep_ret].values
        scaler = StandardScaler()
        Xs = scaler.fit_transform(X)
        try:
            lasso = LassoCV(cv=5, max_iter=20000, random_state=42).fit(Xs, y)
        except Exception as e:
            print(f"  [SKIP] {crude_label}: {e}"); continue
        kept = sum(abs(c) > 1e-8 for c in lasso.coef_)
        print(f"  {crude_label:5s}  alpha*={lasso.alpha_:.4f}  kept {kept}/{len(regs)} regressors")
        for r, c in zip(regs, lasso.coef_):
            rows.append({
                "crude": crude_label, "regressor": r,
                "std_coef": round(float(c), 5),
                "kept": "yes" if abs(c) > 1e-8 else "DROPPED",
                "abs_std_coef": round(abs(float(c)), 5),
            })

    if not rows:
        return
    tab = pd.DataFrame(rows)
    ws = wb.create_sheet("Lasso_pruning")
    ws.append(["Lasso feature pruning (advisory) -- Returns models"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    notes = [
        "Standardised LassoCV on the RETURNS family (Δlog price ~ Δfactors).",
        "Variables with |std_coef| = 0 are candidates to drop from the spec.",
        "The production OLS model is unchanged; this is a diagnostic only.",
        "Use this to motivate a simplified spec if multiple regressors are",
        "consistently dropped across both Light and Heavy.",
    ]
    for n in notes:
        ws.append([n])
    ws.append([])
    ws.append(["Lasso coefficients (standardised)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for r in dataframe_to_rows(tab, index=False, header=True):
        ws.append(r)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col), default=10
        ) + 2


def run_all() -> None:
    df = pd.read_csv(config.FEATURES_CSV, index_col="date", parse_dates=True)
    wb = Workbook()
    wb.remove(wb.active)
    summary_rows: list = []

    _fit_levels_family(df, wb, summary_rows)
    _fit_returns_family(df, wb, summary_rows)
    _fit_differential(df, wb, summary_rows)
    _write_cointegration_sheet(df, wb)
    _write_lasso_sheet(df, wb)

    summary = pd.DataFrame(summary_rows)
    ws = wb.create_sheet("Summary", 0)
    ws.append(["OLS Oil Price Models — Summary (3 families)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Notes"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    notes = [
        "Three model families:",
        "  LEVELS       - log(real price); high R^2 driven mostly by AR(1) lag.",
        "  RETURNS      - Δlog(real price); honest fit, much lower R^2.",
        "  DIFFERENTIAL - WCS-WTI ($/bbl); Enbridge-relevant heavy/light spread.",
        "Inference: Newey-West HAC standard errors, maxlags=6.",
        f"Real prices in {config.REAL_PRICE_BASE_YEAR} dollars; CPI = FRED CPIAUCSL.",
        f"Long-term war = active conflict lasting >= {config.LONG_WAR_MONTHS} months.",
        "Out-of-sample validation: see backtest.py and output/backtest_results.csv.",
        "",
        "WHICH MODEL TO TRUST: the LEVELS R^2 is misleading (mostly the lag).",
        "The RETURNS model is the honest one. The DIFFERENTIAL model is the",
        "structural one for Enbridge. Always cross-check with the backtest.",
    ]
    for n in notes: ws.append([n])
    ws.append([])
    ws.append(["Model fit summary (in-sample)"]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col), default=10
        ) + 2

    ws = wb.create_sheet("Data_recent")
    recent = df.tail(60).reset_index()
    for r in dataframe_to_rows(recent, index=False, header=True):
        ws.append(r)

    wb.save(config.RESULTS_XLSX)
    print(f"\nResults saved: {config.RESULTS_XLSX}")


if __name__ == "__main__":
    run_all()
